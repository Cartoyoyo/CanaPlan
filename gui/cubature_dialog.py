# -*- coding: utf-8 -*-
import csv
import os

from qgis.PyQt.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QFileDialog, QMessageBox, QLabel, QApplication,
    QRadioButton, QButtonGroup, QCheckBox, QGroupBox, QDialogButtonBox,
)
from qgis.core import QgsProject

from ..tools import i18n
from qgis.PyQt.QtCore import Qt, QUrl
from qgis.PyQt.QtGui import QColor, QDesktopServices
from ..tools import errlog


def _libelle_type(valeur):
    """Traduit la valeur du champ type pour l'affichage seulement.

    'Conduite' et 'Branchement' servent de clés de filtrage dans les
    résultats : elles ne doivent pas être traduites à la source.
    """
    if valeur == 'Conduite':
        return i18n.tr('rap_vol_conduite')
    if valeur == 'Branchement':
        return i18n.tr('col_branchement')
    return valeur or '—'


def _libelle_reseau(code, prefixe=False):
    """« Réseau EU — Eaux Usées », traduit. prefixe ajoute le mot Réseau."""
    libelle = i18n.tr('rap_eaux_usees' if code == 'EU' else 'rap_eaux_pluviales')
    if prefixe:
        return i18n.tr('rap_reseau_titre', code=code, libelle=libelle)
    return "%s — %s" % (code, libelle)


def _nb(valeur, decimales=2):
    """Nombre formaté selon la langue, pour l'affichage à l'écran."""
    return i18n.nombre(valeur, decimales)


def _params_segments(config, separateur=" — "):
    """Segments « libellé = valeur » des paramètres de calcul, traduits.

    Servent tels quels à l'écran, au PDF et au XLSX : un seul endroit à
    corriger quand un paramètre change.
    """
    cfg = config

    def _avec_mat(texte, cle_mat):
        mat = cfg.get(cle_mat)
        return f"{texte}{separateur}{mat}" if mat else texte

    segments = [
        _avec_mat("%s = %s m" % (i18n.tr('rap_par_lit_pose'),
                                 _nb(cfg.get('ep_lit_pose', 0.10))),
                  'materiau_lit_pose'),
        "%s = %s m" % (i18n.tr('rap_par_larg_cond', reseau='EU'),
                       _nb(cfg.get('largeur_conduite_eu', 0.80))),
        "%s = %s m" % (i18n.tr('rap_par_larg_cond', reseau='EP'),
                       _nb(cfg.get('largeur_conduite_ep', 0.80))),
        "%s = %s m" % (i18n.tr('rap_par_larg_branch', reseau='EU'),
                       _nb(cfg.get('largeur_branchement_eu', 0.60))),
        "%s = %s m" % (i18n.tr('rap_par_larg_branch', reseau='EP'),
                       _nb(cfg.get('largeur_branchement_ep', 0.60))),
    ]
    return segments


def _params_remblai_segments(config, separateur=" — "):
    """Segments propres au remblai et à la chaussée, traduits."""
    cfg = config
    segments = [
        (("%s : %s m" % (i18n.tr('rap_par_enrobage'),
                         _nb(cfg.get('ep_enrobage', 0.15))))
         + (f"{separateur}{cfg['materiau_enrobage']}"
            if cfg.get('materiau_enrobage') else "")),
        "%s : %s" % (i18n.tr('rap_par_remblai'),
                     cfg.get('materiau_remblai', '0/31.5')),
    ]
    if cfg.get('chaussee_inf', False):
        segments.append(
            ("%s : %s m" % (i18n.tr('rap_par_ch_inf'),
                            _nb(cfg.get('ep_chaussee_inf', 0.20))))
            + (f"{separateur}{cfg['materiau_chaussee_inf']}"
               if cfg.get('materiau_chaussee_inf') else ""))
    if cfg.get('chaussee_sup', False):
        segments.append(
            ("%s : %s m" % (i18n.tr('rap_par_ch_sup'),
                            _nb(cfg.get('ep_chaussee_sup', 0.08))))
            + (f"{separateur}{cfg['materiau_chaussee_sup']}"
               if cfg.get('materiau_chaussee_sup') else ""))
    return segments


class CubatureDialog(QDialog):

    # Clés i18n des colonnes, dans l'ordre d'affichage. La liste sert aussi de
    # référence de longueur (len) pour les fusions et les sous-totaux.
    _COLUMNS = [
        'col_id', 'col_reseau', 'col_type', 'col_materiau', 'col_diametre_court',
        'col_nom_debut', 'col_nom_fin',
        'col_long_2d', 'col_long_3d', 'col_pente',
        'col_profondeur_moy', 'col_largeur', 'rap_surf_ouv',
        'col_deblai',
        'rap_recap_lit_pose', 'rap_recap_enrobage', 'rap_recap_conduite',
        'rap_recap_ch_inf', 'rap_recap_ch_sup', 'rap_recap_remblai',
    ]

    def _column_labels(self):
        """En-têtes traduits, dans la langue courante."""
        return [i18n.tr(cle) for cle in self._COLUMNS]

    # Colonnes de détail du remblai : n'existent qu'en mode show_remblai,
    # et certaines sont en plus conditionnées à un flag de config (chaussée).
    _VOL_BREAKDOWN = [
        dict(key='lit_pose', field='vol_lit_pose',
             pdf_label='rap_vol_lit_pose', recap_label='rap_recap_lit_pose', pdf_width=9, recap_width=20),
        dict(key='enrobage', field='vol_enrobage',
             pdf_label='rap_vol_enrobage', recap_label='rap_recap_enrobage', pdf_width=9, recap_width=20),
        dict(key='conduite', field='vol_conduite',
             pdf_label='rap_vol_conduite', recap_label='rap_recap_conduite', pdf_width=9, recap_width=20),
        dict(key='ch_inf', field='vol_chaussee_inf',
             pdf_label='rap_vol_ch_inf', recap_label='rap_recap_ch_inf', pdf_width=8, recap_width=18,
             cfg_flag='chaussee_inf'),
        dict(key='ch_sup', field='vol_chaussee_sup',
             pdf_label='rap_vol_ch_sup', recap_label='rap_recap_ch_sup', pdf_width=8, recap_width=18,
             cfg_flag='chaussee_sup'),
        dict(key='remblai', field='vol_remblai',
             pdf_label='rap_vol_remblai', recap_label='rap_recap_remblai', pdf_width=10, recap_width=22),
    ]

    def _active_vol_cols(self):
        """Colonnes de détail du remblai actives pour ce rapport (vide en mode cubature simple)."""
        if not self.show_remblai:
            return []
        return [c for c in self._VOL_BREAKDOWN
                if not c.get('cfg_flag') or self.config.get(c['cfg_flag'], False)]

    @staticmethod
    def _avg(values):
        """Moyenne des valeurs non None, ou None si aucune valeur valide."""
        vals = [v for v in values if v is not None]
        return sum(vals) / len(vals) if vals else None

    @staticmethod
    def _listing_ouvrages(noeuds):
        """Listing nominatif : [(nom, TN, profondeur)], trié par nom.

        Tri naturel : REU2 doit précéder REU10, ce qu'un tri alphabétique
        ferait à l'envers.
        """
        import re

        def cle(nom):
            return [int(m) if m.isdigit() else m.lower()
                    for m in re.split(r'(\d+)', nom)]

        return [(nom, tn, prof)
                for nom, (tn, prof) in sorted(noeuds.items(),
                                              key=lambda kv: cle(kv[0]))]

    def _synthese_data(self):
        """Données de synthèse par réseau (EU/EP) : tronçons et branchements
        groupés par matériau + diamètre (nb + linéaire), et comptage des
        regards/tabourets référencés dans le rapport en cours."""
        def _groupes(items):
            groups = {}
            for r in items:
                key = (r.get('materiau') or '—', r.get('diametre'))
                g = groups.setdefault(key, {'cnt': 0, 'long': 0.0})
                g['cnt'] += 1
                g['long'] += r.get('l3d') or 0.0
            return sorted(groups.items(), key=lambda kv: (kv[0][0], -(kv[0][1] or 0)))

        def _profondeurs_noeuds(items, avec_debut):
            """{nom d'ouvrage: (TN, profondeur la plus grande relevée)}.

            Un regard partagé par deux tronçons est vu deux fois ; on garde la
            profondeur la plus grande, celle qui commande la hauteur d'élément.
            """
            noeuds = {}
            for r in items:
                paires = [(r.get('nom_fin'), r.get('tn_fin'), r.get('prof_fin'))]
                if avec_debut:
                    paires.append((r.get('nom_debut'), r.get('tn_debut'),
                                   r.get('prof_debut')))
                for nom, tn, prof in paires:
                    if not nom:
                        continue
                    ancien_tn, ancienne_prof = noeuds.get(nom, (None, None))
                    if ancienne_prof is None or (prof is not None
                                                 and prof > ancienne_prof):
                        noeuds[nom] = (tn if tn is not None else ancien_tn, prof)
                    elif ancien_tn is None and tn is not None:
                        noeuds[nom] = (tn, ancienne_prof)
            return noeuds

        data = []
        for r_name in ('EU', 'EP'):
            reseau_results = [r for r in self.results if r.get('reseau') == r_name]
            if not reseau_results:
                continue

            troncons = [r for r in reseau_results if r.get('type') == 'Conduite']
            branchements = [r for r in reseau_results if r.get('type') == 'Branchement']

            # Profondeur de chaque ouvrage, relevée sur les conduites qui s'y
            # raccordent : prof_debut au nœud amont, prof_fin au nœud aval.
            # Un regard partagé par deux tronçons est vu deux fois ; on garde
            # la valeur la plus profonde, qui commande la hauteur d'élément.
            regards = _profondeurs_noeuds(troncons, avec_debut=True)
            # Le nœud amont d'un branchement est un piquage sur la conduite
            # mère, pas un ouvrage : seul le nœud aval est un tabouret.
            tabourets = _profondeurs_noeuds(branchements, avec_debut=False)

            data.append(dict(
                reseau=r_name,
                troncons_groupes=_groupes(troncons),
                troncons_total=(len(troncons), sum(r.get('l3d') or 0.0 for r in troncons)),
                branchements_groupes=_groupes(branchements),
                branchements_total=(len(branchements), sum(r.get('l3d') or 0.0 for r in branchements)),
                nb_regards=len(regards),
                nb_tabourets=len(tabourets),
                regards_listing=self._listing_ouvrages(regards),
                tabourets_listing=self._listing_ouvrages(tabourets),
            ))
        return data

    def __init__(self, results, config, parent=None, bfs_prefix=None, show_remblai=False):
        super().__init__(parent)
        self.results = results
        self.config = config
        self.bfs_prefix = bfs_prefix  # ex: "REP01_REP04" pour le mode BFS
        self._ouvrir_dossier = True   # False le temps d'un export groupé
        self.show_remblai = show_remblai
        self.setWindowTitle(i18n.tr('cb_titre'))
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowMinimizeButtonHint
            | Qt.WindowMaximizeButtonHint
            | Qt.WindowSystemMenuHint
        )
        self.setMinimumSize(1000, 450)
        self.setAttribute(Qt.WA_DeleteOnClose, False)

        self._build_ui()
        self._update_info_label()
        self._populate_table()
        self._adjust_size_to_content()

    def _adjust_size_to_content(self):
        """Redimensionne la fenêtre pour montrer un maximum de lignes et de
        colonnes visibles sans scroll, dans la limite de l'écran disponible.
        Appelé à l'ouverture et à chaque bascule du détail remblai (le nombre
        de colonnes visibles change alors)."""
        self.table.resizeRowsToContents()
        self.table.resizeColumnsToContents()

        header_h = self.table.horizontalHeader().height()
        rows_h = sum(self.table.rowHeight(i) for i in range(self.table.rowCount()))
        frame_h = 2 * self.table.frameWidth()
        chrome_h = (
            self.info.sizeHint().height()
            + self.cb_remblai.sizeHint().height()
            + self.total_label.sizeHint().height()
            + 90  # boutons d'export + marges de layout
        )
        ideal_h = header_h + rows_h + frame_h + chrome_h

        cols_w = sum(self.table.columnWidth(c) for c in range(self.table.columnCount())
                     if not self.table.isColumnHidden(c))
        v_scrollbar_w = self.table.verticalScrollBar().sizeHint().width()
        frame_w = 2 * self.table.frameWidth()
        ideal_w = cols_w + v_scrollbar_w + frame_w + 30  # marges de layout

        screen = QApplication.primaryScreen()
        avail = screen.availableGeometry() if screen else None
        max_h = int(avail.height() * 0.9) if avail else 900
        max_w = int(avail.width() * 0.9) if avail else 1400

        width = min(max(ideal_w, 1000), max_w)
        height = min(max(ideal_h, 450), max_h)
        self.resize(width, height)

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        # Info config
        self.info = QLabel()
        self.info.setWordWrap(True)
        self.info.setStyleSheet("color: #555; font-size: 10px; padding: 2px;")
        layout.addWidget(self.info)

        # Option remblai
        self.cb_remblai = QCheckBox(
            i18n.tr('cb_detail_remblai'))
        self.cb_remblai.setChecked(self.show_remblai)
        self.cb_remblai.toggled.connect(self._on_toggle_remblai)
        layout.addWidget(self.cb_remblai)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(len(self._COLUMNS))
        self.table.setHorizontalHeaderLabels(self._column_labels())
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.table, 1)

        # Total
        self.total_label = QLabel()
        self.total_label.setStyleSheet(
            "font-weight: bold; font-size: 13px; padding: 6px;"
        )
        layout.addWidget(self.total_label)

        # Boutons export
        btn_row = QHBoxLayout()
        btn_row.addStretch()

        btn_csv = QPushButton(i18n.tr('cb_export_csv'))
        btn_csv.clicked.connect(self._export_csv)
        btn_row.addWidget(btn_csv)

        btn_pdf = QPushButton(i18n.tr('cb_export_pdf'))
        btn_pdf.clicked.connect(self._export_pdf)
        btn_row.addWidget(btn_pdf)

        try:
            import openpyxl
            btn_xlsx = QPushButton(i18n.tr('cb_export_xlsx'))
            btn_xlsx.clicked.connect(self._export_xlsx)
            btn_row.addWidget(btn_xlsx)
        except ImportError as _err:
            errlog.ignored(_err, "cubature_dialog._build_ui:341")

        layout.addLayout(btn_row)

    def _on_toggle_remblai(self, checked):
        self.show_remblai = checked
        self._update_info_label()
        self._populate_table()
        self._adjust_size_to_content()

    def _update_info_label(self):
        segments = _params_segments(self.config)
        if self.show_remblai:
            segments += _params_remblai_segments(self.config)
        self.info.setText("  |  ".join(segments))

    def _populate_table(self):
        eu_results = [r for r in self.results if r.get('reseau') == 'EU']
        ep_results = [r for r in self.results if r.get('reseau') == 'EP']

        self.table.setRowCount(0)
        row = 0

        total_volume = 0.0
        total_surface = 0.0
        total_count = 0

        for reseau_name, reseau_results, color_hex in [
            ('EU', eu_results, '#CC0000'),
            ('EP', ep_results, '#0044CC'),
        ]:
            if not reseau_results:
                continue

            # Ligne de groupe
            self.table.insertRow(row)
            item = QTableWidgetItem(_libelle_reseau(reseau_name, prefixe=True))
            item.setBackground(QColor(color_hex))
            item.setForeground(QColor('white'))
            font = item.font()
            font.setBold(True)
            item.setFont(font)
            self.table.setItem(row, 0, item)
            self.table.setSpan(row, 0, 1, len(self._COLUMNS))
            row += 1

            sous_total = 0.0
            sous_surface = 0.0
            sous_l2d = 0.0
            sous_l3d = 0.0
            sous_vol = {c['key']: 0.0 for c in self._VOL_BREAKDOWN}
            for r in reseau_results:
                self.table.insertRow(row)
                self._set_row(row, r)
                if r.get('volume') is not None:
                    sous_total += r['volume']
                if r.get('surface') is not None:
                    sous_surface += r['surface']
                sous_l2d += r.get('l2d') or 0.0
                sous_l3d += r.get('l3d') or 0.0
                for c in self._VOL_BREAKDOWN:
                    sous_vol[c['key']] += r.get(c['field']) or 0.0
                row += 1

            # Ligne sous-total : label sur ID..Nom fin, puis un sous-total
            # par colonne numérique (m, m², m³).
            self.table.insertRow(row)
            color_lighter = QColor(color_hex).lighter(185)
            label_item = QTableWidgetItem(i18n.tr('rap_sous_total_court', reseau=reseau_name))
            label_item.setBackground(color_lighter)
            font = label_item.font()
            font.setBold(True)
            label_item.setFont(font)
            self.table.setItem(row, 0, label_item)
            self.table.setSpan(row, 0, 1, 7)  # ID, Réseau, Type, Matériau, Ø, Nom début, Nom fin

            subtotal_vals = {7: sous_l2d, 8: sous_l3d, 12: sous_surface}
            for i, c in enumerate(self._VOL_BREAKDOWN):
                subtotal_vals[13 + i] = sous_vol[c['key']]
            subtotal_vals[len(self._COLUMNS) - 1] = sous_total

            for col in range(1, len(self._COLUMNS)):
                text = f"{subtotal_vals[col]:.2f}" if col in subtotal_vals else ""
                cell = QTableWidgetItem(text)
                cell.setFont(font)
                cell.setBackground(color_lighter)
                if text:
                    cell.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, col, cell)
            row += 1

            total_volume += sous_total
            total_surface += sous_surface
            total_count += len(reseau_results)

        label_text = i18n.tr('cb_total_ligne', nb=total_count,
                                surface=_nb(total_surface),
                                deblai=_nb(total_volume))
        self.total_label.setText(label_text)
        self.table.resizeColumnsToContents()

        # Afficher/masquer les colonnes remblai (Vol. lit pose .. Vol. remblai)
        # selon la case à cocher, en réévaluant l'état à chaque appel : il faut
        # explicitement les ré-afficher quand show_remblai repasse à True, sinon
        # elles restent masquées depuis le dernier appel avec show_remblai=False.
        active_keys = {c['key'] for c in self._active_vol_cols()}
        for i, c in enumerate(self._VOL_BREAKDOWN):
            self.table.setColumnHidden(13 + i, c['key'] not in active_keys)

    def _set_row(self, row, data):
        diam = data.get('diametre')
        vals = [
            str(data.get('id', '—')),
            data.get('reseau', ''),
            data.get('type', '—'),
            data.get('materiau', '—') or '—',
            f"{diam:.0f}" if diam is not None else '—',
            data.get('nom_debut', '—'),
            data.get('nom_fin', '—'),
            f"{data.get('l2d', 0):.2f}",
            f"{data.get('l3d', 0):.2f}",
            f"{data.get('pente_pct', 0):.2f}" if data.get('pente_pct') is not None else '—',
            f"{data.get('prof_moy', 0):.2f}" if data.get('prof_moy') is not None else i18n.tr('rap_absence_fe'),
            f"{data.get('largeur', 0):.2f}",
            f"{data.get('surface', 0):.2f}" if data.get('surface') is not None else '—',
            f"{data.get('vol_lit_pose', 0):.2f}" if data.get('vol_lit_pose') is not None else '—',
            f"{data.get('vol_enrobage', 0):.2f}" if data.get('vol_enrobage') is not None else '—',
            f"{data.get('vol_conduite', 0):.2f}" if data.get('vol_conduite') is not None else '—',
            f"{data.get('vol_chaussee_inf', 0):.2f}" if data.get('vol_chaussee_inf') is not None else '—',
            f"{data.get('vol_chaussee_sup', 0):.2f}" if data.get('vol_chaussee_sup') is not None else '—',
            f"{data.get('vol_remblai', 0):.2f}" if data.get('vol_remblai') is not None else '—',
            f"{data.get('volume', 0):.2f}" if data.get('volume') is not None else '—',
        ]
        grey = data.get('err_debut') or data.get('err_fin')
        for col, val in enumerate(vals):
            item = QTableWidgetItem(val)
            item.setTextAlignment(Qt.AlignCenter if col >= 7 else Qt.AlignLeft | Qt.AlignVCenter)
            if grey:
                item.setForeground(QColor('#999999'))
            self.table.setItem(row, col, item)

    # ------------------------------------------------------------------ exports

    def exporter_fichiers(self, out_dir, pdf=False, xlsx=False, csv=False):
        """Écrit les exports demandés dans out_dir, sans aucune boîte de dialogue.

        Utilisé par l'export groupé, où le dialogue n'est jamais affiché.
        Retourne la liste des chemins réellement écrits.
        """
        formats = (
            (pdf,  "pdf",  self._export_pdf),
            (xlsx, "xlsx", self._export_xlsx),
            (csv,  "csv",  self._export_csv),
        )
        ecrits = []
        self._ouvrir_dossier = False
        try:
            for actif, ext, exporter in formats:
                if not actif:
                    continue
                path = os.path.join(out_dir, self._default_filename(ext))
                exporter(path=path)
                if os.path.exists(path):
                    ecrits.append(path)
        finally:
            self._ouvrir_dossier = True
        return ecrits

    def _open_folder(self, path):
        # En export groupé plusieurs fichiers partent d'affilée : ouvrir
        # l'explorateur à chaque écriture serait insupportable.
        if not self._ouvrir_dossier:
            return
        folder = os.path.dirname(os.path.abspath(path))
        QDesktopServices.openUrl(QUrl.fromLocalFile(folder))

    def _default_filename(self, ext):
        # Tous les fichiers produits par l'outil sont prefixes "cubature"
        # pour rester groupes dans l'explorateur.
        parts = ["cubature"]
        if self.show_remblai:
            parts.append("remblai")
        if self.bfs_prefix:
            parts += [self.bfs_prefix, "tranchee"]
        else:
            parts.append("tranchees")
        return "_".join(parts) + "." + ext

    def _export_csv(self, checked=False, path=None):
        if path is None:
            path, _ = QFileDialog.getSaveFileName(
                self, i18n.tr('cb_enregistrer_csv'),
                self._default_filename("csv"), i18n.tr('fic_csv'))
        if not path:
            return
        try:
            vol_cols = self._active_vol_cols()
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                header = self._column_labels()[:13]
                header.append(i18n.tr('rap_deblai_total'))
                header += [i18n.tr('col_csv_remblai',
                                   libelle=i18n.tr(c['recap_label']))
                           for c in vol_cols]
                writer.writerow(header)

                for r in self.results:
                    row = [
                        r.get('id'), r.get('reseau'), r.get('type'),
                        r.get('materiau'),
                        r.get('diametre') if r.get('diametre') is not None else '',
                        r.get('nom_debut'), r.get('nom_fin'),
                        r.get('l2d'), r.get('l3d'),
                        r.get('pente_pct') if r.get('pente_pct') is not None else '',
                        r.get('prof_moy') if r.get('prof_moy') is not None else '',
                        r.get('largeur'),
                        r.get('surface') if r.get('surface') is not None else '',
                    ]
                    row.append(r.get('volume') if r.get('volume') is not None else '')
                    row += [r.get(c['field']) if r.get(c['field']) is not None else '' for c in vol_cols]
                    writer.writerow(row)
            self._open_folder(path)
        except Exception as e:
            QMessageBox.critical(self, i18n.tr('cb_err_csv'), str(e))

    def _export_pdf(self, checked=False, path=None):
        if path is None:
            path, _ = QFileDialog.getSaveFileName(
                self, i18n.tr('cb_enregistrer_pdf'),
                self._default_filename("pdf"), i18n.tr('fic_pdf'))
        if not path:
            return
        try:
            self._export_pdf_reportlab(path)
        except ImportError:
            try:
                # Installation a la demande de reportlab. argv est entierement
                # constant : l'interpreteur courant et des litteraux. Pas de
                # shell, aucune valeur saisie par l'utilisateur.
                import subprocess, sys  # nosec B404
                QApplication.setOverrideCursor(Qt.WaitCursor)
                subprocess.check_call(  # nosec B603
                    [sys.executable, "-m", "pip", "install", "reportlab"],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                QApplication.restoreOverrideCursor()
                self._export_pdf_reportlab(path)
            except Exception as e:
                QApplication.restoreOverrideCursor()
                QMessageBox.critical(
                    self, i18n.tr('ct_export_pdf_titre'),
                    fi18n.tr('cb_reportlab', erreur=e))
        except Exception as e:
            QMessageBox.critical(self, i18n.tr('cb_err_pdf'), str(e))

    # Bornes de largeur des colonnes du tableau de détail, en mm. Le minimum
    # garde une colonne cliquable même vide ; le maximum empêche qu'un nom
    # d'ouvrage à rallonge mange toute la page.
    _COL_MIN_MM = 8
    _COL_MAX_MM = 34

    @staticmethod
    def _texte_brut(rows):
        """Grille de texte nu à partir de lignes de Paragraph ou de chaînes.

        _auto_col_widths mesure du texte ; le récapitulatif, lui, est déjà
        construit en Paragraph balisés.
        """
        import re
        grille = []
        for ligne in rows:
            brute = []
            for cellule in ligne:
                texte = getattr(cellule, 'text', cellule) or ''
                texte = re.sub(r'<[^>]+>', '', str(texte))
                brute.append(texte.replace('&nbsp;', ' '))
            grille.append(brute)
        return grille

    @staticmethod
    def _auto_col_widths(raw_data, avail_width):
        """Largeurs de colonnes déduites du texte réellement présent.

        Les largeurs étaient auparavant figées en dur, ce qui coupait
        « Branchement » en « Branche / ment » et « -24.19 » en « -24.1 / 9 ».
        On mesure ici chaque cellule avec la police qui la rendra, on borne,
        puis on répartit l'espace restant — ou on réduit proportionnellement
        si le total déborde.
        """
        from reportlab.pdfbase import pdfmetrics
        from reportlab.lib.units import mm

        if not raw_data:
            return None
        n_cols = len(raw_data[0])
        mini, maxi = CubatureDialog._COL_MIN_MM * mm, CubatureDialog._COL_MAX_MM * mm
        # padding horizontal appliqué par le TableStyle (3+3), plus 4 pt de
        # garde : sans marge, un budget saturé à 100 % fait replier le texte
        # pour quelques centièmes de point.
        marge = 10

        largeurs = []
        for col in range(n_cols):
            besoin = 0.0
            for ligne_idx, ligne in enumerate(raw_data):
                if col >= len(ligne):
                    continue
                texte = str(ligne[col] or '')
                if not texte:
                    continue
                # l'en-tête est en gras et un point plus grand
                police, taille = ('Helvetica-Bold', 7) if ligne_idx == 0                     else ('Helvetica', 7)
                besoin = max(besoin,
                             pdfmetrics.stringWidth(texte, police, taille))
            largeurs.append(min(max(besoin + marge, mini), maxi))

        total = sum(largeurs)
        if total > avail_width:
            facteur = avail_width / total
            return [w * facteur for w in largeurs]
        # Espace disponible en trop : on le donne aux colonnes de texte, qui
        # sont les seules à en tirer profit.
        rab = avail_width - total
        if rab > 0:
            extensibles = [i for i, w in enumerate(largeurs) if w < maxi]
            if extensibles:
                part = rab / len(extensibles)
                for i in extensibles:
                    largeurs[i] += part
        return largeurs

    def _export_pdf_reportlab(self, path):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm, mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
            KeepTogether, Image,
        )
        from ..tools import i18n

        def _n(valeur, decimales=2, vide='—'):
            """Nombre formaté selon la langue (virgule ou point décimal)."""
            return i18n.nombre(valeur, decimales, vide)
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus.flowables import HRFlowable
        from datetime import date

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title2', parent=styles['Title'],
            fontSize=16, spaceAfter=2*mm, textColor=colors.HexColor('#1a1a1a'),
        )
        subtitle_style = ParagraphStyle(
            'SubTitle2', parent=styles['Normal'],
            fontSize=9, textColor=colors.HexColor('#555555'), spaceAfter=1*mm,
        )
        h3_style = ParagraphStyle(
            'H3', parent=styles['Heading3'],
            fontSize=11, spaceBefore=4*mm, spaceAfter=2*mm,
        )
        param_style = ParagraphStyle(
            'Param', parent=styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#444444'),
        )
        footer_style = ParagraphStyle(
            'Footer', parent=styles['Normal'],
            fontSize=7, textColor=colors.HexColor('#999999'),
        )
        cell_style = ParagraphStyle(
            'Cell', parent=styles['Normal'],
            fontSize=7, leading=9, alignment=TA_CENTER,
        )
        cell_left = ParagraphStyle(
            'CellLeft', parent=cell_style, alignment=TA_LEFT,
        )
        # Le libellé de sous-total est collé à droite, contre ses valeurs.
        cell_right = ParagraphStyle(
            'CellRight', parent=cell_style, alignment=TA_RIGHT,
        )
        # Un Paragraph ignore le TEXTCOLOR du TableStyle : la couleur de
        # l'en-tête doit être portée par le style lui-même.
        entete_style = ParagraphStyle(
            'Entete', parent=cell_style, fontName='Helvetica-Bold',
            fontSize=7, leading=8, textColor=colors.white,
        )
        total_style = ParagraphStyle(
            'Total', parent=styles['Normal'],
            fontSize=10, textColor=colors.HexColor('#222222'),
            alignment=TA_RIGHT,
        )

        # Projet
        projet = QgsProject.instance()
        projet_nom = i18n.tr('rap_projet_defaut')
        # Le nom qui compte est celui du projet CanaPlan (.bet) ; le projet
        # QGIS ne sert que de repli quand rien n'a encore été enregistré.
        try:
            from ..tools.projet_bet import current_bet_name
            nom_bet = current_bet_name()
        except Exception:
            nom_bet = ""
        if nom_bet:
            projet_nom = nom_bet
        elif projet:
            base = projet.baseName() or projet.fileName()
            if base:
                projet_nom = base

        # Orientation paysage pour le mode remblai (plus de colonnes)
        page_size = landscape(A4) if self.show_remblai else A4
        # Largeur utile pour les tableaux, marges déduites
        avail_width = page_size[0] - 30*mm
        doc = SimpleDocTemplate(
            path, pagesize=page_size,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=12*mm, bottomMargin=15*mm,
        )
        story = []

        # ── En-tête ──────────────────────────────────────────────
        report_type = i18n.tr('rap_remblai') if self.show_remblai \
            else i18n.tr('rap_cubature')
        # Le logo est le PNG à fond transparent ; ReportLab ne lit pas le SVG
        # sans svglib, absent des installations QGIS.
        logo_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "icon", "logo-full.png")
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=48*mm, height=12.6*mm)
            logo.hAlign = 'LEFT'
            story.append(logo)
            story.append(Spacer(1, 2*mm))
        # Le titre porte le nom du chantier, pas celui du plugin.
        story.append(Paragraph(f"{projet_nom} — {report_type}", title_style))
        story.append(Paragraph(
            i18n.tr('rap_date', date=date.today().strftime('%d/%m/%Y')),
            subtitle_style,
        ))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#CC0000')))
        story.append(Spacer(1, 3*mm))

        # ── Paramètres ───────────────────────────────────────────
        story.append(Paragraph(i18n.tr('rap_parametres'), h3_style))
        cfg = self.config
        sep = " &nbsp;|&nbsp; "
        params_text = (sep.join(_params_segments(cfg, separateur=" "))
                       + "<br/>"
                       + sep.join(_params_remblai_segments(cfg, separateur=" ")))
        story.append(Paragraph(params_text, param_style))
        story.append(Spacer(1, 4*mm))

        # ── Tableaux par réseau ──────────────────────────────────
        eu_results = [r for r in self.results if r.get('reseau') == 'EU']
        ep_results = [r for r in self.results if r.get('reseau') == 'EP']

        # Styles pour sous-groupes
        sousgroupe_style = ParagraphStyle(
            'SousGroupe', parent=styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#444444'),
            spaceBefore=3*mm, spaceAfter=1*mm,
            leftIndent=2*mm,
        )
        total_reseau_style = ParagraphStyle(
            'TotalReseau', parent=styles['Normal'],
            fontSize=9, textColor=colors.HexColor('#222222'),
            alignment=TA_RIGHT, spaceBefore=2*mm, spaceAfter=1*mm,
        )

        recap_data = []  # liste de (reseau, total, count, color, {type: (count, vol)})
        grand_total = 0.0
        grand_count = 0

        for reseau_name, reseau_results, color_hex, color_light in [
            ('EU', eu_results, '#CC0000', '#FFECEC'),
            ('EP', ep_results, '#0044CC', '#ECF0FF'),
        ]:
            if not reseau_results:
                continue

            story.append(Paragraph(
                _libelle_reseau(reseau_name, prefixe=True),
                ParagraphStyle(
                    'ReseauTitle', parent=styles['Heading3'],
                    fontSize=10, textColor=colors.HexColor(color_hex),
                    spaceBefore=5*mm, spaceAfter=2*mm,
                ),
            ))

            vol_cols = self._active_vol_cols()
            # Les largeurs ne sont plus figées : elles se déduisent du
            # contenu réel (voir _auto_col_widths). Des libellés traduits plus
            # longs — l'allemand notamment — ne peuvent donc plus déborder.
            # (libellé, unité) : l'unité passe sur une deuxième ligne de
            # l'en-tête. Sur une seule ligne, « Prof. moy. (m) » mesure 55 pt
            # pour une valeur de 12 pt, et les 13 colonnes ne tenaient plus
            # dans la largeur utile de l'A4 portrait.
            pdf_cols = [("ID", ""), (i18n.tr('col_type'), ""),
                        (i18n.tr('rap_mat'), ""),
                        ("Ø", "mm"), (i18n.tr('rap_debut'), ""),
                        (i18n.tr('rap_fin'), ""),
                        (i18n.tr('rap_l2d'), "m"), (i18n.tr('rap_l3d'), "m"),
                        (i18n.tr('rap_pente'), "%"),
                        (i18n.tr('rap_prof_moy'), "m"),
                        (i18n.tr('rap_larg'), "m"), (i18n.tr('rap_surf'), "m²")]
            # Ordre du chantier : on métré, on déblaie, puis on remblaie.
            # Le déblai précède donc sa décomposition en remblai.
            IDX_DEBLAI = len(pdf_cols)          # 12
            pdf_cols.append((i18n.tr('rap_deblai'), "m³"))
            idx_vol_debut = len(pdf_cols)       # 13
            pdf_cols += [(i18n.tr(c['pdf_label']), "m³") for c in vol_cols]
            n_total = len(pdf_cols)  # nombre total de colonnes
            entetes_cells = [
                Paragraph(
                    lib + (("<br/><font size=5.5>(%s)</font>" % uni) if uni else ""),
                    entete_style)
                for lib, uni in pdf_cols
            ]
            # Bandeau de groupe : les colonnes de remblai décomposent le
            # déblai, il faut que ça se voie. Colonnes 12..n-2 = remblai,
            # dernière colonne = déblai.
            bandeau = [''] * n_total
            if vol_cols:
                bandeau[0] = Paragraph(i18n.tr('rap_metre'), entete_style)
                bandeau[IDX_DEBLAI] = Paragraph(
                    i18n.tr('rap_deblai') + "<br/><font size=5.5>(m³)</font>",
                    entete_style)
                # Le déblai tient sur les deux lignes d'en-tête : la cellule
                # du dessous lui est fusionnée.
                entetes_cells[IDX_DEBLAI] = Paragraph("", entete_style)
                bandeau[idx_vol_debut] = Paragraph(
                    i18n.tr('rap_remblai_decomposition'), entete_style)
            # Deux lignes de mesure : chaque ligne de l'en-tête est mesurée
            # séparément, sinon on additionnerait libellé et unité.
            mesure_entetes = [[lib for lib, _u in pdf_cols],
                              [("(%s)" % u if u else "") for _l, u in pdf_cols]]
            # Colonnes sommables / moyennables dans la ligne de sous-total.
            # Une pente ou une largeur ne s'additionnent pas : on en donne la
            # moyenne, signalée en italique.
            IDX_SOMME = [6, 7, 11, IDX_DEBLAI] + list(range(idx_vol_debut, n_total))
            # La pente (indice 8) est volontairement absente : moyenner des
            # pentes de tronçons de longueurs différentes n'a pas de sens, et
            # la fenêtre Cubature des tranchées ne la totalise pas non plus.
            IDX_MOYENNE = [9, 10]
            CHAMPS = {6: 'l2d', 7: 'l3d', 8: 'pente_pct', 9: 'prof_moy',
                      10: 'largeur', 11: 'surface', IDX_DEBLAI: 'volume'}
            for _i, _c in enumerate(vol_cols):
                CHAMPS[idx_vol_debut + _i] = _c['field']
            header_color = colors.HexColor(color_hex)
            cell_bg = colors.HexColor(color_light)

            reseau_total = 0.0
            reseau_surface = 0.0
            reseau_count = 0
            reseau_detail = {}  # {type: (count, vol, bd, surface)}

            for sous_type, sous_label in [
                    ('Conduite', i18n.tr('cb_conduites')),
                    ('Branchement', i18n.tr('cb_branchements'))]:
                sous_results = [r for r in reseau_results if r.get('type') == sous_type]
                if not sous_results:
                    reseau_detail[sous_type] = (0, 0.0, None, 0.0)
                    continue

                sous_titre = Paragraph(f"▸ {sous_label}", sousgroupe_style)

                # texte brut en parallele des Paragraph : sert au calcul
                # automatique des largeurs de colonnes
                raw_data = list(mesure_entetes)
                table_data = [entetes_cells]
                if vol_cols:
                    # Le bandeau passe en première ligne, l'en-tête détaillé
                    # devient la seconde.
                    table_data.insert(0, list(bandeau))
                sous_total = 0.0
                sous_surface = 0.0
                sous_bd = None  # breakdown des volumes
                for r in sous_results:
                    diam = r.get('diametre')
                    raw_row = [
                        str(r.get('id', '—')),
                        _libelle_type(r.get('type')),
                        r.get('materiau', '—') or '—',
                        _n(diam, 0),
                        r.get('nom_debut', '—'),
                        r.get('nom_fin', '—'),
                        _n(r.get('l2d', 0)),
                        _n(r.get('l3d', 0)),
                        _n(r.get('pente_pct')),
                        _n(r.get('prof_moy'), vide=i18n.tr('rap_absence_fe_court')),
                        _n(r.get('largeur', 0)),
                        _n(r.get('surface')),
                    ]
                    raw_row.append(_n(r.get('volume')))
                    raw_row += [_n(r.get(c['field'])) for c in vol_cols]
                    raw_data.append(raw_row)
                    table_data.append(
                        [Paragraph(v, cell_style) for v in raw_row])
                    if r.get('volume') is not None:
                        sous_total += r['volume']
                        if r.get('surface') is not None:
                            sous_surface += r['surface']
                        if vol_cols:
                            if sous_bd is None:
                                sous_bd = {c['key']: 0.0 for c in vol_cols}
                            for c in vol_cols:
                                sous_bd[c['key']] += r.get(c['field'], 0) or 0.0

                # Ligne sous-total : toutes les colonnes additionnables sont
                # totalisées, les autres (pente, profondeur, largeur) reçoivent
                # leur moyenne, en italique pour ne pas être lues comme un total.
                sous_raw = [''] * n_total
                libelle_st = i18n.tr('rap_sous_total', libelle=sous_label,
                                     reseau=reseau_name,
                                     nb=len(sous_results))
                sous_raw[0] = libelle_st
                sous_cells = [Paragraph("<b>%s</b>" % libelle_st, cell_right)] \
                    + [''] * (n_total - 1)
                for idx in IDX_SOMME:
                    champ = CHAMPS.get(idx)
                    if champ is None:
                        continue
                    total_col = sum(r.get(champ) or 0.0 for r in sous_results)
                    sous_raw[idx] = _n(total_col)
                    sous_cells[idx] = Paragraph(
                        "<b>%s</b>" % sous_raw[idx], cell_style)
                for idx in IDX_MOYENNE:
                    champ = CHAMPS.get(idx)
                    valeurs = [r.get(champ) for r in sous_results
                               if r.get(champ) is not None]
                    if not valeurs:
                        continue
                    sous_raw[idx] = _n(sum(valeurs) / len(valeurs))
                    sous_cells[idx] = Paragraph(
                        "<i>%s</i>" % sous_raw[idx], cell_style)
                raw_data.append(sous_raw)
                table_data.append(sous_cells)
                reseau_total += sous_total
                reseau_surface += sous_surface
                reseau_count += len(sous_results)

                col_widths = self._auto_col_widths(raw_data, avail_width)
                n_entetes = 2 if vol_cols else 1
                tbl = Table(table_data, colWidths=col_widths,
                            repeatRows=n_entetes)
                # derniere ligne d'en-tete (0 ou 1 selon la presence du bandeau)
                der = n_entetes - 1
                cmds = [
                    ('BACKGROUND', (0, 0), (-1, der), header_color),
                    ('TEXTCOLOR', (0, 0), (-1, der), colors.white),
                    ('FONTSIZE', (0, 0), (-1, der), 7),
                    ('FONTNAME', (0, 0), (-1, der), 'Helvetica-Bold'),
                    ('ALIGN', (0, 0), (-1, der), 'CENTER'),
                    ('BOTTOMPADDING', (0, 0), (-1, der), 4),
                    ('BACKGROUND', (0, n_entetes), (-1, -2), cell_bg),
                    ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
                    ('ROWBACKGROUNDS', (0, n_entetes), (-1, -2),
                     [cell_bg, colors.white]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EEEEEE')),
                    ('LINEABOVE', (0, -1), (-1, -1), 1, header_color),
                    # Le libellé du sous-total occupe les colonnes
                    # d'identification, sinon ReportLab le casse lettre par
                    # lettre dans la colonne ID large de 9 mm.
                    ('SPAN', (0, -1), (5, -1)),
                ]
                if vol_cols:
                    # Bandeau : « Métré » sur les colonnes descriptives,
                    # « Remblai » sur sa décomposition, « Total » sur le déblai.
                    cmds += [
                        ('SPAN', (0, 0), (IDX_DEBLAI - 1, 0)),
                        ('SPAN', (IDX_DEBLAI, 0), (IDX_DEBLAI, 1)),
                        ('SPAN', (idx_vol_debut, 0), (n_total - 1, 0)),
                        # Traits verticaux marquant les deux frontières
                        ('LINEBEFORE', (IDX_DEBLAI, 0), (IDX_DEBLAI, -1),
                         1.2, colors.white),
                        ('LINEBEFORE', (idx_vol_debut, 0), (idx_vol_debut, -1),
                         1.2, colors.white),
                    ]
                tbl.setStyle(TableStyle(cmds))
                # Titre et tableau restent solidaires : évite une page qui
                # s'ouvre sur une ligne de sous-total orpheline.
                story.append(KeepTogether([sous_titre, tbl]))
                story.append(Spacer(1, 1*mm))
                reseau_detail[sous_type] = (len(sous_results), sous_total, sous_bd, sous_surface)

            # Total réseau
            story.append(Paragraph(
                i18n.tr('rap_total_reseau', reseau=reseau_name,
                        surface=_n(reseau_surface),
                        deblai=_n(reseau_total), nb=reseau_count),
                total_reseau_style,
            ))

            recap_data.append((reseau_name, reseau_total, reseau_count, color_hex, reseau_detail, reseau_surface))
            grand_total += reseau_total
            grand_count += reseau_count

        # ── Récapitulatif projet ─────────────────────────────────
        story.append(Spacer(1, 5*mm))
        story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#333333')))
        story.append(Spacer(1, 3*mm))
        recap_titre = Paragraph(i18n.tr('rap_recapitulatif'), ParagraphStyle(
            'RecapTitle', parent=styles['Heading2'],
            fontSize=12, spaceAfter=3*mm, textColor=colors.HexColor('#222222'),
        ))

        vol_cols = self._active_vol_cols()
        vol_keys = [c['key'] for c in vol_cols]

        # Meme ordre que le tableau de detail : metre, deblai, remblai.
        recap_cols = ["", i18n.tr('rap_nb'),
                      i18n.tr('col_profondeur_moy'),
                      i18n.tr('col_largeur'),
                      i18n.tr('rap_surf_ouv'),
                      i18n.tr('rap_deblai') + " (m³)"] + \
            [i18n.tr(c['recap_label']) for c in vol_cols]
        if vol_cols:
            recap_widths = [46*mm, 10*mm, 15*mm, 13*mm, 17*mm, 20*mm] + \
                [c['recap_width']*mm for c in vol_cols]
        else:
            recap_widths = [70*mm, 15*mm, 18*mm, 15*mm, 25*mm, 27*mm]
        n_recap = len(recap_cols)

        recap_rows = [recap_cols]
        sous_total_rows = []
        grand_bd = {k: 0.0 for k in vol_keys}

        for r_name, r_vol, r_count, r_color, r_detail, r_surf in recap_data:
            reseau_label = _libelle_reseau(r_name)
            recap_rows.append([Paragraph(f"<b>{reseau_label}</b>", cell_left)] + [''] * (n_recap - 1))

            reseau_bd = {k: 0.0 for k in vol_keys}
            for sous_type, sous_label in [
                    ('Conduite', i18n.tr('cb_conduites')),
                    ('Branchement', i18n.tr('cb_branchements'))]:
                cnt, vol, bd, surf = r_detail.get(sous_type, (0, 0.0, None, 0.0))
                sous_results = [r for r in self.results
                                if r.get('reseau') == r_name and r.get('type') == sous_type]
                profmoy_avg = self._avg([r.get('prof_moy') for r in sous_results])
                largeur_avg = self._avg([r.get('largeur') for r in sous_results])
                row_data = [Paragraph(f"&nbsp;&nbsp;&nbsp;{sous_label}", cell_left),
                            Paragraph(str(cnt), cell_style),
                            Paragraph(_n(profmoy_avg), cell_style),
                            Paragraph(_n(largeur_avg), cell_style)]
                row_data.append(Paragraph(_n(surf) if cnt > 0 else '—', cell_style))
                row_data.append(Paragraph(_n(vol) if cnt > 0 else '—', cell_style))
                for key in vol_keys:
                    val = bd.get(key, 0.0) if bd else 0.0
                    row_data.append(Paragraph(_n(val) if (cnt > 0 and bd) else '—', cell_style))
                    if bd:
                        reseau_bd[key] += val
                recap_rows.append(row_data)

            # Sous-total réseau
            reseau_results_r = [r for r in self.results if r.get('reseau') == r_name]
            reseau_profmoy_avg = self._avg([r.get('prof_moy') for r in reseau_results_r])
            reseau_largeur_avg = self._avg([r.get('largeur') for r in reseau_results_r])
            st_row = [Paragraph("<i>%s</i>" % i18n.tr('rap_sous_total_court', reseau=r_name),
                      cell_left),
                      Paragraph(f"<b>{r_count}</b>", cell_style),
                      Paragraph("<b>%s</b>" % _n(reseau_profmoy_avg), cell_style),
                      Paragraph("<b>%s</b>" % _n(reseau_largeur_avg), cell_style)]
            st_row.append(Paragraph("<b>%s</b>" % _n(r_surf), cell_style))
            st_row.append(Paragraph("<b>%s</b>" % _n(r_vol), cell_style))
            for key in vol_keys:
                st_row.append(Paragraph("<b>%s</b>" % _n(reseau_bd[key]), cell_style))
                grand_bd[key] += reseau_bd[key]
            recap_rows.append(st_row)
            sous_total_rows.append(len(recap_rows) - 1)

        # TOTAL PROJET
        grand_surf = sum(r[5] for r in recap_data)  # reseau_surface
        grand_profmoy_avg = self._avg([r.get('prof_moy') for r in self.results])
        grand_largeur_avg = self._avg([r.get('largeur') for r in self.results])
        total_row_data = [Paragraph("<b>%s</b>" % i18n.tr('rap_total_projet'), cell_left),
                          Paragraph(f"<b>{grand_count}</b>", cell_style),
                          Paragraph("<b>%s</b>" % _n(grand_profmoy_avg), cell_style),
                          Paragraph("<b>%s</b>" % _n(grand_largeur_avg), cell_style)]
        total_row_data.append(Paragraph("<b>%s</b>" % _n(grand_surf), cell_style))
        total_row_data.append(Paragraph("<b>%s</b>" % _n(grand_total), cell_style))
        for key in vol_keys:
            total_row_data.append(Paragraph("<b>%s</b>" % _n(grand_bd[key]), cell_style))
        recap_rows.append(total_row_data)
        total_row = len(recap_rows) - 1

        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7 if vol_cols else 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('BACKGROUND', (0, total_row), (-1, total_row), colors.HexColor('#E8E8E8')),
            ('LINEABOVE', (0, total_row), (-1, total_row), 1.5, colors.HexColor('#333333')),
        ]
        for st_row in sous_total_rows:
            style_cmds.append(('BACKGROUND', (0, st_row), (-1, st_row), colors.HexColor('#EEEEEE')))
            style_cmds.append(('LINEABOVE', (0, st_row), (-1, st_row), 0.8, colors.HexColor('#666666')))

        recap_widths = self._auto_col_widths(
            self._texte_brut(recap_rows), avail_width)
        # repeatRows : si le tableau doit malgré tout se scinder, la ligne
        # d'en-tête est réimprimée en haut de la page suivante.
        recap_tbl = Table(recap_rows, colWidths=recap_widths, repeatRows=1)
        recap_tbl.setStyle(TableStyle(style_cmds))

        # Titre et tableau solidaires : le récapitulatif tenait sur une page,
        # mais son en-tête restait seul en bas de la précédente.
        story.append(KeepTogether([recap_titre, recap_tbl]))

        # ── Synthèse des ouvrages ──────────────────────────────────
        # Pas de saut de page : la synthèse tient sous le récapitulatif et
        # occupait sinon une page au quart remplie.
        story.append(Spacer(1, 6*mm))
        synth_titre = Paragraph(i18n.tr('rap_synthese'), ParagraphStyle(
            'SynthTitle', parent=styles['Heading2'],
            fontSize=12, spaceAfter=3*mm, textColor=colors.HexColor('#222222'),
        ))

        def _synth_style(couleur):
            """Style commun des tableaux de synthèse, à la couleur du réseau."""
            return [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(couleur)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EEEEEE')),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor(couleur)),
            ]

        def _synth_table(groupes, total, label_total, couleur):
            rows = [[i18n.tr('col_materiau'), "Ø (mm)",
                     i18n.tr('rap_long_totale')]]
            for (mat, diam), g in groupes:
                rows.append([
                    Paragraph(mat, cell_left),
                    Paragraph(_n(diam, 0), cell_style),
                    Paragraph(_n(g['long']), cell_style),
                ])
            _cnt_total, long_total = total
            rows.append([
                Paragraph(f"<b>{label_total}</b>", cell_left), '',
                Paragraph("<b>%s</b>" % _n(long_total), cell_style),
            ])
            tbl = Table(rows, colWidths=[55*mm, 25*mm, 35*mm])
            tbl.setStyle(TableStyle(_synth_style(couleur)))
            return tbl

        def _ouvrages_table(listing, label, couleur):
            """Listing des regards ou tabourets : TN et profondeur.

            TN et profondeur viennent des conduites raccordées (nœud amont et
            nœud aval) ; ils n'étaient pas exploités alors qu'ils commandent la
            hauteur des éléments à commander. Le sous-total donne le nombre.
            """
            rows = [[i18n.tr('col_ouvrage'), i18n.tr('col_tn'),
                     i18n.tr('col_profondeur')]]
            for nom, tn, prof in listing:
                rows.append([
                    Paragraph(nom, cell_left),
                    Paragraph(_n(tn), cell_style),
                    Paragraph(_n(prof), cell_style),
                ])
            rows.append([
                Paragraph("<b>%s</b>" % label, cell_left),
                Paragraph(f"<b>{len(listing)}</b>", cell_style), '',
            ])
            tbl = Table(rows, colWidths=[55*mm, 25*mm, 35*mm], repeatRows=1)
            tbl.setStyle(TableStyle(_synth_style(couleur)))
            return tbl

        premier_reseau = True
        for d in self._synthese_data():
            synth_color = '#CC0000' if d['reseau'] == 'EU' else '#0044CC'
            reseau_label = _libelle_reseau(d['reseau'])
            bloc = []
            if premier_reseau:
                bloc.append(synth_titre)
                premier_reseau = False
            bloc.append(Paragraph(reseau_label, ParagraphStyle(
                'SynthReseau', parent=styles['Heading3'],
                fontSize=10, textColor=colors.HexColor(synth_color),
                spaceBefore=4*mm, spaceAfter=2*mm,
            )))
            bloc.append(Paragraph("▸ " + i18n.tr('rap_troncons'), sousgroupe_style))
            bloc.append(_synth_table(d['troncons_groupes'],
                                     d['troncons_total'], i18n.tr('rap_total_troncons'),
                                     synth_color))
            bloc.append(Spacer(1, 2*mm))
            bloc.append(Paragraph("▸ " + i18n.tr('cb_branchements'), sousgroupe_style))
            bloc.append(_synth_table(d['branchements_groupes'],
                                     d['branchements_total'],
                                     i18n.tr('rap_total_branchements'), synth_color))
            if d['regards_listing']:
                bloc.append(Spacer(1, 2*mm))
                bloc.append(Paragraph("▸ " + i18n.tr('rap_regards'), sousgroupe_style))
                bloc.append(_ouvrages_table(d['regards_listing'],
                                            i18n.tr('rap_total_regards'),
                                            synth_color))
            if d['tabourets_listing']:
                bloc.append(Spacer(1, 2*mm))
                bloc.append(Paragraph("▸ " + i18n.tr('rap_tabourets'), sousgroupe_style))
                bloc.append(_ouvrages_table(d['tabourets_listing'],
                                            i18n.tr('rap_total_tabourets'),
                                            synth_color))
            bloc.append(Spacer(1, 5*mm))
            # Un réseau ne se scinde pas au milieu d'un de ses tableaux.
            story.append(KeepTogether(bloc))

        # ── Pied de page ─────────────────────────────────────────
        def add_page_number(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.HexColor('#999999'))
            canvas.drawString(15*mm, 10*mm,
                              "%s — %s" % (projet_nom, report_type))
            canvas.drawRightString(page_size[0] - 15*mm, 10*mm, i18n.tr('rap_page', n=canvas.getPageNumber()))
            canvas.restoreState()

        doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        self._open_folder(path)

    def _export_xlsx(self, checked=False, path=None):
        if path is None:
            path, _ = QFileDialog.getSaveFileName(
                self, i18n.tr('cb_enregistrer_xlsx'),
                self._default_filename("xlsx"), i18n.tr('fic_xlsx'))
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # ── Feuille 1 : Récapitulatif ─────────────────────────
            ws_recap = wb.active
            ws_recap.title = i18n.tr('rap_recapitulatif_court')

            title_font = Font(bold=True, size=14, color='1a1a1a')
            h2_font = Font(bold=True, size=11, color='333333')
            header_font = Font(bold=True, color='FFFFFF', size=10)
            header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
            bold_font = Font(bold=True, size=10)
            normal_font = Font(size=10)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )
            bottom_border = Border(
                bottom=Side(style='medium'),
            )
            subtotal_fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
            total_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            red_fill = PatternFill(start_color='FFECEC', end_color='FFECEC', fill_type='solid')
            blue_fill = PatternFill(start_color='ECF0FF', end_color='ECF0FF', fill_type='solid')

            def apply_border(ws, row, cols, extra_border=None):
                for c in range(1, cols + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.border = thin_border

            # Titre
            report_type = i18n.tr('rap_remblai') if self.show_remblai \
                else i18n.tr('rap_cubature')
            ws_recap.merge_cells('A1:D1')
            c = ws_recap.cell(row=1, column=1, value=i18n.tr(
                'rap_projet_type', projet=i18n.tr('rap_projet_defaut'),
                type=report_type))
            c.font = title_font
            c.alignment = Alignment(horizontal='left')

            from datetime import date
            ws_recap.merge_cells('A2:D2')
            projet = QgsProject.instance()
            projet_nom = i18n.tr('rap_projet_defaut')
            if projet:
                base = projet.baseName() or projet.fileName()
                if base:
                    projet_nom = base
            c = ws_recap.cell(row=2, column=1,
                value=i18n.tr('rap_projet_date', projet=projet_nom,
                              date=date.today().strftime('%d/%m/%Y')))
            c.font = Font(size=9, color='555555')

            # Paramètres
            ws_recap.merge_cells('A4:D4')
            c = ws_recap.cell(row=4, column=1, value=i18n.tr('rap_parametres'))
            c.font = h2_font
            cfg = self.config
            params = "  |  ".join(_params_segments(cfg)
                                  + _params_remblai_segments(cfg))
            c = ws_recap.cell(row=5, column=1, value=params)
            c.font = Font(size=8, color='444444')

            # Tableau récap
            vol_cols = self._active_vol_cols()
            vol_keys = [c['key'] for c in vol_cols]
            vol_field = {c['key']: c['field'] for c in vol_cols}

            recap_cols = ["", i18n.tr('rap_nb'),
                          i18n.tr('col_profondeur_moy'),
                          i18n.tr('col_largeur'),
                          i18n.tr('rap_surf_ouv'),
                          i18n.tr('rap_deblai') + " (m³)"] + \
                [i18n.tr(c['recap_label']) for c in vol_cols]
            n_recap = len(recap_cols)

            recap_header_row = 7
            for col_idx, col_name in enumerate(recap_cols, 1):
                cell = ws_recap.cell(row=recap_header_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            eu_results = [r for r in self.results if r.get('reseau') == 'EU']
            ep_results = [r for r in self.results if r.get('reseau') == 'EP']
            grand_total = 0.0
            grand_surface = 0.0
            grand_count = 0
            grand_bd = {k: 0.0 for k in vol_keys}

            row = recap_header_row + 1
            for r_name, reseau_results, color_hex in [
                ('EU', eu_results, 'CC0000'),
                ('EP', ep_results, '0044CC'),
            ]:
                fill = red_fill if r_name == 'EU' else blue_fill
                reseau_label = _libelle_reseau(r_name)
                ws_recap.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_recap)
                c = ws_recap.cell(row=row, column=1, value=reseau_label)
                c.font = Font(bold=True, size=10)
                c.fill = fill
                for cc in range(1, n_recap + 1):
                    ws_recap.cell(row=row, column=cc).border = thin_border
                row += 1

                reseau_total = 0.0
                reseau_surface = 0.0
                reseau_count = 0
                reseau_bd = {k: 0.0 for k in vol_keys}

                for sous_type, sous_label in [
                    ('Conduite', i18n.tr('cb_conduites')),
                    ('Branchement', i18n.tr('cb_branchements'))]:
                    sous_results = [r for r in reseau_results if r.get('type') == sous_type]
                    cnt = len(sous_results)
                    vol = sum(r.get('volume', 0.0) or 0.0 for r in sous_results)
                    surf = sum(r.get('surface', 0.0) or 0.0 for r in sous_results)
                    profmoy_avg = self._avg([r.get('prof_moy') for r in sous_results])
                    largeur_avg = self._avg([r.get('largeur') for r in sous_results])

                    c = ws_recap.cell(row=row, column=1, value=f"   {sous_label}")
                    c.font = normal_font; c.border = thin_border
                    c = ws_recap.cell(row=row, column=2, value=cnt)
                    c.font = normal_font; c.alignment = Alignment(horizontal='center'); c.border = thin_border
                    c = ws_recap.cell(row=row, column=3, value=round(profmoy_avg, 2) if profmoy_avg is not None else '—')
                    c.font = normal_font; c.alignment = Alignment(horizontal='center'); c.border = thin_border
                    c = ws_recap.cell(row=row, column=4, value=round(largeur_avg, 2) if largeur_avg is not None else '—')
                    c.font = normal_font; c.alignment = Alignment(horizontal='center'); c.border = thin_border

                    col = 5
                    c = ws_recap.cell(row=row, column=col, value=round(surf, 2) if cnt > 0 else '—'); col += 1
                    c.font = normal_font; c.alignment = Alignment(horizontal='center'); c.border = thin_border
                    c = ws_recap.cell(row=row, column=col, value=round(vol, 2) if cnt > 0 else '—'); col += 1
                    c.font = normal_font; c.alignment = Alignment(horizontal='center'); c.border = thin_border
                    for key in vol_keys:
                        v = sum(r.get(vol_field[key], 0.0) or 0.0 for r in sous_results) if cnt > 0 else None
                        c = ws_recap.cell(row=row, column=col, value=round(v, 2) if v is not None else '—')
                        c.font = normal_font; c.alignment = Alignment(horizontal='center'); c.border = thin_border
                        if v is not None:
                            reseau_bd[key] += v
                        col += 1

                    row += 1
                    reseau_total += vol
                    reseau_surface += surf
                    reseau_count += cnt

                # Sous-total réseau
                for cc in range(1, n_recap + 1):
                    ws_recap.cell(row=row, column=cc).fill = subtotal_fill
                    ws_recap.cell(row=row, column=cc).border = Border(
                        top=Side(style='thin'), bottom=Side(style='thin'),
                        left=Side(style='thin'), right=Side(style='thin'))
                reseau_profmoy_avg = self._avg([r.get('prof_moy') for r in reseau_results])
                reseau_largeur_avg = self._avg([r.get('largeur') for r in reseau_results])
                c = ws_recap.cell(row=row, column=1, value=i18n.tr('rap_sous_total_court', reseau=r_name))
                c.font = Font(bold=True, size=10, italic=True); c.fill = subtotal_fill
                c = ws_recap.cell(row=row, column=2, value=reseau_count)
                c.font = Font(bold=True, size=10); c.fill = subtotal_fill; c.alignment = Alignment(horizontal='center')
                c = ws_recap.cell(row=row, column=3, value=round(reseau_profmoy_avg, 2) if reseau_profmoy_avg is not None else '—')
                c.font = Font(bold=True, size=10); c.fill = subtotal_fill; c.alignment = Alignment(horizontal='center')
                c = ws_recap.cell(row=row, column=4, value=round(reseau_largeur_avg, 2) if reseau_largeur_avg is not None else '—')
                c.font = Font(bold=True, size=10); c.fill = subtotal_fill; c.alignment = Alignment(horizontal='center')

                col = 5
                c = ws_recap.cell(row=row, column=col, value=round(reseau_surface, 2)); col += 1
                c.font = Font(bold=True, size=10); c.fill = subtotal_fill; c.alignment = Alignment(horizontal='center')
                c = ws_recap.cell(row=row, column=col, value=round(reseau_total, 2)); col += 1
                c.font = Font(bold=True, size=10); c.fill = subtotal_fill; c.alignment = Alignment(horizontal='center')
                for key in vol_keys:
                    c = ws_recap.cell(row=row, column=col, value=round(reseau_bd[key], 2))
                    c.font = Font(bold=True, size=10); c.fill = subtotal_fill; c.alignment = Alignment(horizontal='center')
                    grand_bd[key] += reseau_bd[key]
                    col += 1
                row += 1
                grand_total += reseau_total
                grand_surface += reseau_surface
                grand_count += reseau_count

            # TOTAL PROJET
            for cc in range(1, n_recap + 1):
                ws_recap.cell(row=row, column=cc).fill = total_fill
                ws_recap.cell(row=row, column=cc).border = Border(
                    top=Side(style='medium'), bottom=Side(style='thin'),
                    left=Side(style='thin'), right=Side(style='thin'))
            grand_profmoy_avg = self._avg([r.get('prof_moy') for r in self.results])
            grand_largeur_avg = self._avg([r.get('largeur') for r in self.results])
            c = ws_recap.cell(row=row, column=1, value=i18n.tr('rap_total_projet'))
            c.font = Font(bold=True, size=11); c.fill = total_fill
            c = ws_recap.cell(row=row, column=2, value=grand_count)
            c.font = Font(bold=True, size=11); c.fill = total_fill; c.alignment = Alignment(horizontal='center')
            c = ws_recap.cell(row=row, column=3, value=round(grand_profmoy_avg, 2) if grand_profmoy_avg is not None else '—')
            c.font = Font(bold=True, size=11); c.fill = total_fill; c.alignment = Alignment(horizontal='center')
            c = ws_recap.cell(row=row, column=4, value=round(grand_largeur_avg, 2) if grand_largeur_avg is not None else '—')
            c.font = Font(bold=True, size=11); c.fill = total_fill; c.alignment = Alignment(horizontal='center')

            col = 5
            c = ws_recap.cell(row=row, column=col, value=round(grand_surface, 2)); col += 1
            c.font = Font(bold=True, size=11); c.fill = total_fill; c.alignment = Alignment(horizontal='center')
            c = ws_recap.cell(row=row, column=col, value=round(grand_total, 2)); col += 1
            c.font = Font(bold=True, size=11); c.fill = total_fill; c.alignment = Alignment(horizontal='center')
            for key in vol_keys:
                c = ws_recap.cell(row=row, column=col, value=round(grand_bd[key], 2))
                c.font = Font(bold=True, size=11); c.fill = total_fill; c.alignment = Alignment(horizontal='center')
                col += 1

            ws_recap.column_dimensions['A'].width = 40
            if vol_cols:
                for col_idx in range(2, n_recap + 1):
                    ws_recap.column_dimensions[get_column_letter(col_idx)].width = 16
            else:
                ws_recap.column_dimensions['B'].width = 12
                ws_recap.column_dimensions['C'].width = 16
                ws_recap.column_dimensions['D'].width = 14
                ws_recap.column_dimensions['E'].width = 18
                ws_recap.column_dimensions['F'].width = 18

            # ── Feuille 2 : Données détaillées ─────────────────────
            ws_data = wb.create_sheet(i18n.tr('rap_donnees_detaillees'))

            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            bande_fill = PatternFill(start_color='2F5597', end_color='2F5597',
                                     fill_type='solid')

            # Bandeau de groupe, comme dans le PDF : les colonnes N à S
            # décomposent le remblai, la colonne T porte le déblai. Sans lui,
            # « Vol. remblai » et « Déblai » se lisent comme deux totaux de
            # même nature alors que le premier entre dans le second.
            N_COLS = len(self._COLUMNS)          # 20
            COL_DEBLAI = 14                      # N
            COL_VOL_DEBUT, COL_VOL_FIN = 15, N_COLS   # O..T
            for debut, fin, titre in (
                    (1, COL_DEBLAI - 1, i18n.tr('rap_metre')),
                    (COL_DEBLAI, COL_DEBLAI, i18n.tr('col_deblai')),
                    (COL_VOL_DEBUT, COL_VOL_FIN,
                     i18n.tr('rap_remblai_decomposition'))):
                ws_data.merge_cells(start_row=1, start_column=debut,
                                    end_row=1, end_column=fin)
                c = ws_data.cell(row=1, column=debut, value=titre)
                c.font = header_font
                c.alignment = Alignment(horizontal='center')
                for cc in range(debut, fin + 1):
                    ws_data.cell(row=1, column=cc).fill = bande_fill
                    ws_data.cell(row=1, column=cc).border = thin_border

            LIGNE_ENTETE = 2
            for col_idx, col_name in enumerate(self._column_labels(), 1):
                cell = ws_data.cell(row=LIGNE_ENTETE, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            grey_font = Font(color='999999')

            for row_idx, r in enumerate(self.results, LIGNE_ENTETE + 1):
                row_fill = red_fill if r.get('reseau') == 'EU' else blue_fill
                diam = r.get('diametre')
                vals = [
                    r.get('id'), r.get('reseau'), r.get('type'),
                    r.get('materiau') or '',
                    diam if diam is not None else '',
                    r.get('nom_debut'), r.get('nom_fin'),
                    r.get('l2d'), r.get('l3d'),
                    r.get('pente_pct') if r.get('pente_pct') is not None else '',
                    r.get('prof_moy') if r.get('prof_moy') is not None else i18n.tr('rap_absence_fe'),
                    r.get('largeur'),
                    r.get('surface') if r.get('surface') is not None else '',
                    r.get('volume') if r.get('volume') is not None else '',
                    r.get('vol_lit_pose') if r.get('vol_lit_pose') is not None else '',
                    r.get('vol_enrobage') if r.get('vol_enrobage') is not None else '',
                    r.get('vol_conduite') if r.get('vol_conduite') is not None else '',
                    r.get('vol_chaussee_inf') if r.get('vol_chaussee_inf') is not None else '',
                    r.get('vol_chaussee_sup') if r.get('vol_chaussee_sup') is not None else '',
                    r.get('vol_remblai') if r.get('vol_remblai') is not None else '',
                ]
                grey = r.get('err_debut') or r.get('err_fin')
                for col_idx, val in enumerate(vals, 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    if grey:
                        cell.font = grey_font

            for col_idx in range(1, len(self._COLUMNS) + 1):
                ws_data.column_dimensions[get_column_letter(col_idx)].width = 14

            # Activer autofilter sur la feuille données
            premiere_data = LIGNE_ENTETE + 1
            last_data_row = len(self.results) + LIGNE_ENTETE
            ws_data.auto_filter.ref = (
                f"A{LIGNE_ENTETE}:{get_column_letter(N_COLS)}{last_data_row}")
            # En-têtes figés : bandeau et titres restent visibles au défilement.
            ws_data.freeze_panes = f"A{premiere_data}"

            # Masquer colonnes selon mode
            if not self.show_remblai:
                # Colonnes de remblai O..T. La borne s'arretait a R : la
                # colonne « Vol. remblai » restait visible en mode cubature.
                for c in range(COL_VOL_DEBUT, COL_VOL_FIN + 1):
                    ws_data.column_dimensions[get_column_letter(c)].hidden = True
            else:
                if not self.config.get('chaussee_inf', False):
                    ws_data.column_dimensions[get_column_letter(18)].hidden = True
                if not self.config.get('chaussee_sup', False):
                    ws_data.column_dimensions[get_column_letter(19)].hidden = True

            # ── Ligne sous-total avec formules SUBTOTAL ──────────────
            subtotal_row = last_data_row + 2  # saute une ligne
            subtotal_fill_xl = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            subtotal_font = Font(bold=True, size=10)

            # Fusion A-G : "Sous-total lignes affichées"
            ws_data.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=7)
            c = ws_data.cell(row=subtotal_row, column=1, value=i18n.tr('rap_sous_total_affichees'))
            c.font = subtotal_font; c.fill = subtotal_fill_xl
            c.alignment = Alignment(horizontal='right')
            for cc in range(1, 8):
                ws_data.cell(row=subtotal_row, column=cc).fill = subtotal_fill_xl
                ws_data.cell(row=subtotal_row, column=cc).border = Border(
                    top=Side(style='medium'), bottom=Side(style='thin'),
                    left=Side(style='thin'), right=Side(style='thin'))

            # Colonnes numériques avec SUBTOTAL (109=SUM, 101=AVERAGE)
            sum_cols = {8: 'H', 9: 'I', 13: 'M', 14: 'N', 15: 'O', 16: 'P',
                        17: 'Q', 18: 'R', 19: 'S', 20: 'T'}
            # Colonne 10 (J) = Pente : pas de moyenne, cf. remarque dans
            # la construction du sous-total PDF.
            avg_cols = {11: 'K', 12: 'L'}

            for col_idx_1based, col_letter in {**sum_cols, **avg_cols}.items():
                c = ws_data.cell(row=subtotal_row, column=col_idx_1based)
                c.font = subtotal_font; c.fill = subtotal_fill_xl
                c.alignment = Alignment(horizontal='center')
                c.border = Border(
                    top=Side(style='medium'), bottom=Side(style='thin'),
                    left=Side(style='thin'), right=Side(style='thin'))

            for col_idx_1based, col_letter in sum_cols.items():
                ws_data.cell(row=subtotal_row, column=col_idx_1based).value = \
                    f"=SUBTOTAL(109,{col_letter}{premiere_data}:{col_letter}{last_data_row})"

            for col_idx_1based, col_letter in avg_cols.items():
                ws_data.cell(row=subtotal_row, column=col_idx_1based).value = \
                    f"=SUBTOTAL(101,{col_letter}{premiere_data}:{col_letter}{last_data_row})"

            # ── Feuille 3 : Synthèse ouvrages (tronçons/branchements par Ø/matériau, regards, tabourets) ──
            ws_synth = wb.create_sheet(i18n.tr('rap_synthese_court'))
            ws_synth.merge_cells('A1:C1')
            c = ws_synth.cell(row=1, column=1, value=i18n.tr('rap_synthese'))
            c.font = title_font
            c.alignment = Alignment(horizontal='left')

            def _write_groupes_table(row_s, groupes, total, label_total):
                for col_idx, h in enumerate([i18n.tr('col_materiau'),
                                             i18n.tr('col_diametre_court'),
                                             i18n.tr('rap_long_totale')], 1):
                    cell = ws_synth.cell(row=row_s, column=col_idx, value=h)
                    cell.font = header_font; cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
                row_s += 1
                for (mat, diam), g in groupes:
                    ws_synth.cell(row=row_s, column=1, value=mat).border = thin_border
                    c = ws_synth.cell(row=row_s, column=2, value=round(diam, 1) if diam is not None else '—')
                    c.alignment = Alignment(horizontal='center'); c.border = thin_border
                    c = ws_synth.cell(row=row_s, column=3, value=round(g['long'], 2))
                    c.alignment = Alignment(horizontal='center'); c.border = thin_border
                    row_s += 1
                _cnt_total, long_total = total
                c = ws_synth.cell(row=row_s, column=1, value=label_total)
                c.font = Font(bold=True); c.border = thin_border
                ws_synth.cell(row=row_s, column=2).border = thin_border
                c = ws_synth.cell(row=row_s, column=3, value=round(long_total, 2))
                c.font = Font(bold=True); c.alignment = Alignment(horizontal='center'); c.border = thin_border
                return row_s + 2

            row_s = 3
            for d in self._synthese_data():
                fill = red_fill if d['reseau'] == 'EU' else blue_fill
                reseau_label = _libelle_reseau(d['reseau'])

                ws_synth.merge_cells(start_row=row_s, start_column=1, end_row=row_s, end_column=3)
                c = ws_synth.cell(row=row_s, column=1, value=reseau_label)
                c.font = Font(bold=True, size=11)
                c.fill = fill
                for cc in range(1, 4):
                    ws_synth.cell(row=row_s, column=cc).border = thin_border
                row_s += 1

                row_s = _write_groupes_table(row_s, d['troncons_groupes'], d['troncons_total'],
                                              i18n.tr('rap_total_troncons'))
                row_s = _write_groupes_table(row_s, d['branchements_groupes'], d['branchements_total'],
                                              i18n.tr('rap_total_branchements'))

                def _write_listing(row_s, listing, titre, label_total):
                    """Listing nominatif des ouvrages : TN et profondeur.

                    Remplace le simple comptage : la hauteur des éléments à
                    commander se lit ouvrage par ouvrage.
                    """
                    if not listing:
                        return row_s
                    c = ws_synth.cell(row=row_s, column=1, value=titre)
                    c.font = Font(bold=True, size=10)
                    row_s += 1
                    for col_idx, h in enumerate(
                            [i18n.tr('col_ouvrage'), i18n.tr('col_tn'),
                     i18n.tr('col_profondeur')], 1):
                        cell = ws_synth.cell(row=row_s, column=col_idx, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = thin_border
                    row_s += 1
                    for nom, tn, prof in listing:
                        ws_synth.cell(row=row_s, column=1, value=nom).border = thin_border
                        c = ws_synth.cell(row=row_s, column=2, value=tn if tn is not None else '—')
                        c.alignment = Alignment(horizontal='center'); c.border = thin_border
                        c = ws_synth.cell(row=row_s, column=3, value=prof if prof is not None else '—')
                        c.alignment = Alignment(horizontal='center'); c.border = thin_border
                        row_s += 1
                    c = ws_synth.cell(row=row_s, column=1, value=label_total)
                    c.font = Font(bold=True); c.border = thin_border
                    ws_synth.cell(row=row_s, column=2).border = thin_border
                    c = ws_synth.cell(row=row_s, column=3, value=len(listing))
                    c.font = Font(bold=True); c.alignment = Alignment(horizontal='center')
                    c.border = thin_border
                    return row_s + 2

                row_s = _write_listing(row_s, d['regards_listing'],
                                       i18n.tr('rap_regards'),
                                       i18n.tr('rap_total_regards'))
                row_s = _write_listing(row_s, d['tabourets_listing'],
                                       i18n.tr('rap_tabourets'),
                                       i18n.tr('rap_total_tabourets'))

                row_s += 2

            ws_synth.column_dimensions['A'].width = 22
            ws_synth.column_dimensions['B'].width = 12
            ws_synth.column_dimensions['C'].width = 18

            wb.save(path)
            self._open_folder(path)
        except Exception as e:
            QMessageBox.critical(self, i18n.tr('cb_err_xlsx'), str(e))


class CubatureOptionsDialog(QDialog):
    """Dialogue d'accueil : périmètre, types d'ouvrages, mode BFS ou axe."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(i18n.tr('cb_options_titre'))
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # ── Périmètre ────────────────────────────────────────────
        group_perim = QGroupBox(i18n.tr('cb_perimetre'))
        perim_layout = QVBoxLayout()
        self.rb_tout = QRadioButton(i18n.tr('cb_tout'))
        self.rb_eu = QRadioButton(i18n.tr('cb_eu_seul'))
        self.rb_ep = QRadioButton(i18n.tr('cb_ep_seul'))
        self.rb_tout.setChecked(True)
        self._perim_group = QButtonGroup(self)
        self._perim_group.addButton(self.rb_tout, 0)
        self._perim_group.addButton(self.rb_eu, 1)
        self._perim_group.addButton(self.rb_ep, 2)
        perim_layout.addWidget(self.rb_tout)
        perim_layout.addWidget(self.rb_eu)
        perim_layout.addWidget(self.rb_ep)
        group_perim.setLayout(perim_layout)
        layout.addWidget(group_perim)

        # ── Types d'ouvrages ─────────────────────────────────────
        group_types = QGroupBox(i18n.tr('cb_types'))
        types_layout = QVBoxLayout()
        self.cb_conduites = QCheckBox(i18n.tr('cb_conduites'))
        self.cb_conduites.setChecked(True)
        self.cb_branchements = QCheckBox(i18n.tr('cb_branchements'))
        self.cb_branchements.setChecked(True)
        types_layout.addWidget(self.cb_conduites)
        types_layout.addWidget(self.cb_branchements)
        group_types.setLayout(types_layout)
        layout.addWidget(group_types)

        # ── Mode de sélection ─────────────────────────────────────
        group_mode = QGroupBox(i18n.tr('cb_mode'))
        mode_layout = QVBoxLayout()
        self.cb_bfs = QCheckBox(i18n.tr('cb_mode_bfs'))
        self.cb_axe = QCheckBox(i18n.tr('cb_mode_axe'))
        self.cb_bfs.toggled.connect(lambda checked: self.cb_axe.setEnabled(not checked))
        self.cb_axe.toggled.connect(lambda checked: self.cb_bfs.setEnabled(not checked))
        mode_layout.addWidget(self.cb_bfs)
        mode_layout.addWidget(self.cb_axe)
        group_mode.setLayout(mode_layout)
        layout.addWidget(group_mode)

        # ── Boutons ──────────────────────────────────────────────
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._validate_and_accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _validate_and_accept(self):
        if not self.cb_conduites.isChecked() and not self.cb_branchements.isChecked():
            QMessageBox.warning(
                self, i18n.tr('msg_cubature_titre'),
                i18n.tr('cb_choisir_type'))
            return
        self.accept()

    def options(self):
        """Retourne le dict d'options choisies."""
        perim_id = self._perim_group.checkedId()
        perim_map = {0: 'tout', 1: 'EU', 2: 'EP'}
        return {
            'perimetre': perim_map.get(perim_id, 'tout'),
            'conduites': self.cb_conduites.isChecked(),
            'branchements': self.cb_branchements.isChecked(),
            'bfs': self.cb_bfs.isChecked(),
            'axe': self.cb_axe.isChecked(),
        }
